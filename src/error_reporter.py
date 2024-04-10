from openpyxl import load_workbook


def get_expected_result(status_code):
    new_status_code = int(status_code)
    if 400 <= int(new_status_code) < 500:
        return f'Нет ошибки {new_status_code}, контент присутствует.'
    elif int(new_status_code) >= 500:
        return 'Нет php ошибки.'
    else:
        return ''
    
def save_file(work_book, save_path):
    try:
        work_book.save(save_path)
    except PermissionError:
        input(f'Закройте файл {save_path}. Для повтора нажмите Enter')
        save_file(work_book, save_path)
        

def auto_adjust_row_height(work_sheet, length):
    for row in range(1, length):
        work_sheet.row_dimensions[row].height = None
    

def make_save_report_to_excel(issue_data, save_path):
    destination_wb = load_workbook("report_template.xlsx")
    destination_ws = destination_wb.active
    
    for row, issue in zip(destination_ws.iter_rows(min_row=3, max_row=len(issue_data)+2), issue_data):
        row[0].value = f'spider-{issue['id']}'
        row[3].value = f'url: {issue['url']} status code: {issue['status_code']} \nРодитель: {issue['referer']} \nОшибка: {issue['text_error']}'
        row[4].value = f'Скриншот: {issue['screenshot']}'
        row[5].value = get_expected_result(issue['status_code'])
        
    auto_adjust_row_height(destination_ws, len(issue_data) + 4)
    
    save_file(destination_wb, f"projects\\{save_path}.xlsx")
